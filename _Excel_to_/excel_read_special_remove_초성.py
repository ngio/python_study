"""_summary_
    엑셀에서 지정된 열의 2번째 행부터 데이터를 읽어서,
    그 옆 열(column_index + 1)에 특수문자/공백 제거된 문자열을 쓰고,
    그 다음 옆 열(column_index + 2)에 초성만 추출된 문자열을 쓰는 기능
"""

import openpyxl
import re # 정규 표현식 모듈 추가

def process_excel(input_file, column_index):
    # 엑셀 파일 로드
    wb = openpyxl.load_workbook(input_file)
    
    # 활성 워크시트 선택
    sheet = wb.active
    
    # 최대 행 수 가져오기
    max_row = sheet.max_row
    
    print(f"최대 행 수: {sheet.max_row}")
    print(f"읽을 컬럼 인덱스: {column_index} 열")
    
    # 처리된 데이터를 저장할 리스트 (특수문자/공백 제거 버전)
    cleaned_data_list = []
    # 처리된 데이터를 저장할 리스트 (초성 추출 버전)
    initial_consonants_data_list = []
    
    # 2행부터 적용하기 위해 range(2, max_row + 1)
    for row_num in range(2, max_row + 1):
        cell_value = sheet.cell(row=row_num, column=column_index).value
        
        # 셀 값이 None인 경우 빈 문자열로 처리
        if cell_value is None:
            str_value = ""
        else:
            str_value = str(cell_value) # 셀 값을 문자열로 변환
            
        # 1. 특수문자 및 공백 제거
        cleaned_value = remove_special_characters_and_spaces(str_value)
        
        # 2. 제거된 문자열에서 초성 추출
        initial_consonants_only = extract_initial_consonants(cleaned_value)
        
        cleaned_data_list.append(cleaned_value)
        initial_consonants_data_list.append(initial_consonants_only)
        
        print(f"행: {row_num}, 원본 값: '{cell_value}'")
        print(f"  -> 특수문자/공백 제거: '{cleaned_value}'")
        print(f"  -> 초성만 추출: '{initial_consonants_only}'")
        
    # --- 처리된 데이터를 엑셀 파일에 쓰기 ---
    
    # 특수문자/공백 제거된 데이터를 column_index + 1 열에 쓰기
    # original_row_num은 2부터 시작하므로 그대로 활용
    for i, original_row_num in enumerate(range(2, max_row + 1)):
        sheet.cell(row=original_row_num, column=column_index + 1).value = cleaned_data_list[i]
    
    # 초성만 추출된 데이터를 column_index + 2 열에 쓰기
    for i, original_row_num in enumerate(range(2, max_row + 1)):
        sheet.cell(row=original_row_num, column=column_index + 2).value = initial_consonants_data_list[i]
    
    # 변경 사항을 새 Excel 파일로 저장
    output_file = input_file.replace('.xlsx', '_processed_분리.xlsx')
    wb.save(output_file)
    print(f"\n처리된 파일이 '{output_file}'(으)로 저장되었습니다.")


# --- 초성 추출 관련 함수 ---

# 한글 초성 리스트
INITIAL_CONSONANTS = [
    'ㄱ', 'ㄲ', 'ㄴ', 'ㄷ', 'ㄸ', 'ㄹ', 
    'ㅁ', 'ㅂ', 'ㅃ', 'ㅅ', 'ㅆ', 'ㅇ', 
    'ㅈ', 'ㅉ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ'
]

def is_korean_char(ch):
    """문자가 한글 음절인지 확인합니다."""
    return 0xAC00 <= ord(ch) <= 0xD7A3

def get_initial_consonant(ch):
    """한글 음절에서 초성만 추출합니다."""
    if not is_korean_char(ch):
        return '' 
    
    initial_index = (ord(ch) - 0xAC00) // (21 * 28)
    return INITIAL_CONSONANTS[initial_index]

def extract_initial_consonants(text):
    """주어진 텍스트에서 한글 음절의 초성만 추출하여 반환합니다."""
    result = []
    if not isinstance(text, str):
        return ''
        
    for char in text:
        if is_korean_char(char):
            result.append(get_initial_consonant(char))
    return ''.join(result)


# --- 특수문자 및 공백 제거 함수 ---

def remove_special_characters_and_spaces(text):
    """
    주어진 텍스트에서 한글, 영어 대소문자, 숫자 외의 모든 특수문자와 공백을 제거합니다.
    """
    if not isinstance(text, str):
        return ''
        
    # [^가-힣a-zA-Z0-9]는 한글, 영어, 숫자를 제외한 모든 문자
    # re.sub를 사용하여 이들을 빈 문자열로 대체
    cleaned_text = re.sub(r'[^가-힣a-zA-Z0-9]', '', text) 
    return cleaned_text # strip()은 필요 없으므로 제거 (이미 공백이 제거됨)


# --- 사용 예시 ---
input_excel_file = './data/test_01.xlsx' # 여기에 실제 파일 경로를 입력하세요.
column_to_read = 2  # B열 (2번째 열)에서 읽습니다.

# 함수 호출
process_excel(input_excel_file, column_to_read)
