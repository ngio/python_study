# 파이썬 컴파일 경로가 달라서 현재 폴더의 이미지를 호출하지 못할때 작업디렉토리를 변경한다. 
import os
from pathlib import Path
# src 상위 폴더를 실행폴더로 지정하려고 한다.
###real_path = Path(__file__).parent.parent
real_path = Path(__file__).parent
print(real_path)
#작업 디렉토리 변경
os.chdir(real_path) 

"""_summary_
    엑셀에서 1 열을 읽어서 그 열 옆에 데이터 쓰기 
"""

import openpyxl

def process_excel(input_file, column_index):
    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)
    
    # Select the active worksheet
    sheet = wb.active
    
    # Get the maximum row count
    max_row = sheet.max_row
    
    print(" max_row = " + str(sheet.max_row))
    
    # Read the data from the specified column
    column_data = []
    for row in range(1, max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        column_data.append(cell_value)
    
    # Process the data (e.g., modify or analyze it)
    processed_data = [str(item) + " processed" for item in column_data]
    
    # Write the processed data to the next column
    for index, value in enumerate(processed_data, start=1):
        sheet.cell(row = index, column=column_index + 2).value = value
    
    # Save the changes to a new Excel file
    output_file = input_file.replace('.xlsx', '_processed.xlsx')
    wb.save(output_file)

# Usage example
input_excel_file = '/data/test_01.xlsx'  # Replace with the path to your input Excel file
column_to_read = 1  # Column index to read (1 for column A, 2 for column B, etc.)
process_excel(input_excel_file, column_to_read)




